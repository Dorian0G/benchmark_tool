"""
output_generator.py
Writes a three-sheet Excel workbook using openpyxl.
"""

import io
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL  = PatternFill("solid", fgColor="1B3A5C")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="EAF1FB")
NUMBER_FMT   = '#,##0.00'


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert every column to plain Python-native types that openpyxl can
    serialise.  Handles pandas nullable Int64/Float64 extension types where
    <NA> is not the same object as float('nan') and cannot be written to Excel.
    Uses convert_dtypes(dtype_backend='numpy_nullable') to unwrap extension
    arrays, then replaces every remaining NA sentinel with None.
    """
    out = df.copy()
    for col in out.columns:
        out[col] = [
            None if pd.isna(v) else v          # catch pd.NA / np.nan / pd.NaT
            for v in out[col].astype(object)   # unwrap extension array first
        ]
    return out


def _style_header(ws, row_idx: int = 1):
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _alt_rows(ws, start_row: int = 2):
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        if i % 2 == 1:
            for cell in row:
                cell.fill = ALT_ROW_FILL


def _write_df(ws, df: pd.DataFrame, number_cols: list[str] | None = None):
    """Write a DataFrame to a worksheet starting at A1."""
    df = _sanitize_df(df)

    ws.append(df.columns.tolist())
    _style_header(ws)

    number_cols = number_cols or []
    for row in df.itertuples(index=False):
        ws.append(list(row))

    col_indices = {col: df.columns.get_loc(col) + 1 for col in number_cols if col in df.columns}
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in col_indices.values():
                cell.number_format = NUMBER_FMT

    _alt_rows(ws)
    _autowidth(ws)


def generate_excel(
    raw_df: pd.DataFrame,
    clean_df: pd.DataFrame,
    bench_df: pd.DataFrame,
    insights: str = "",
    copilot_prompt: str = "",
    grantees_by_company: dict[str, list[dict]] | None = None,
) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Raw Data"
    _write_df(ws1, raw_df)

    # Cleaned Data sheet uses the post-fill, post-derived dataframe so it
    # matches the "Clean Data" tab in the tool exactly (imputed values + the
    # computed "Giving as % of Revenue" column).
    ws2 = wb.create_sheet("Cleaned Data")
    numeric_cols = [c for c in clean_df.columns if c != "Company"]
    _write_df(ws2, clean_df, number_cols=numeric_cols)

    # Benchmark Summary mirrors the tool's "Benchmark" tab column order:
    # Company | Metric | Data Year | Value | Rank | Industry Average | Percentile
    bench_for_excel = bench_df.copy()
    preferred = ["Company", "Metric", "Data Year", "Value", "Rank", "Industry Average", "Percentile"]
    ordered = [c for c in preferred if c in bench_for_excel.columns]
    extras = [c for c in bench_for_excel.columns if c not in ordered]
    bench_for_excel = bench_for_excel[ordered + extras]

    ws3 = wb.create_sheet("Benchmark Summary")
    _write_df(ws3, bench_for_excel, number_cols=["Value", "Industry Average", "Percentile"])

    # Grantee Directory sheet — one row per grantee, prefixed with the company
    # so all foundations live on a single sortable sheet.
    if grantees_by_company:
        rows: list[dict] = []
        for company, grantees in grantees_by_company.items():
            for g in grantees or []:
                rows.append({
                    "Company":    company,
                    "Grantee":    g.get("grantee", ""),
                    "City":       g.get("city", ""),
                    "State":      g.get("state", ""),
                    "Amount ($)": g.get("amount", 0) or 0,
                    "Purpose":    g.get("purpose", ""),
                    "Year":       g.get("year", ""),
                    "Source":     g.get("source", ""),
                })
        if rows:
            ws_gr = wb.create_sheet("Grantee Directory")
            grantee_df = pd.DataFrame(rows)
            _write_df(ws_gr, grantee_df, number_cols=["Amount ($)"])

    if insights:
        ws4 = wb.create_sheet("AI Insights")
        ws4.append(["AI-Generated Insights"])
        ws4["A1"].font = Font(bold=True, size=13)
        for line in insights.split("\n"):
            ws4.append([line])
        ws4.column_dimensions["A"].width = 90

    if copilot_prompt:
        ws5 = wb.create_sheet("Copilot Prompt")
        ws5.append(["Paste this prompt into copilot.microsoft.com"])
        ws5["A1"].font = Font(bold=True, size=12, color="0078D4")
        ws5.append([""])
        for line in copilot_prompt.split("\n"):
            ws5.append([line])
        ws5.column_dimensions["A"].width = 100

    wb.properties.description = (
        f"Generated by Utility Benchmark Tool · {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(
    raw_df: pd.DataFrame,
    clean_df: pd.DataFrame,
    bench_df: pd.DataFrame,
    insights: str = "",
    path: str | Path = "outputs/benchmark.xlsx",
) -> Path:
    data = generate_excel(raw_df, clean_df, bench_df, insights)
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_bytes(data)
    logger.info("Excel saved to %s", p)
    return p
